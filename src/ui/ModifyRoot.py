'''
@coding=UTF-8
@auther:Yan Chen, Fei JianLin, Yang Cheng, Tang YuHao
@time:2020-06
@ui:Modify Info of Root
'''

# 导入模块
import tkinter as tk
from tkinter import messagebox
from tkinter.ttk import *
import openpyxl

'''
下面是生成查看学生信息页面的代码
'''
def ModifyRoot():
    modifyRoot = tk.Tk()
    modifyRoot.title('欢迎来到学生管理系统！')
    modifyRoot.geometry('600x350+300+300')
    modifyRoot['bg'] = 'lightblue'
    modifyRoot.attributes('-alpha', 0.9)

    # 点击"提交"按钮的操作
    def clickSubR():
        data = openpyxl.load_workbook('../lib/Students.xlsx')
        sheetnames = data.get_sheet_names()
        table = data.get_sheet_by_name(sheetnames[0])
        table = data.active
        nrows = table.max_row                               # 获得最大行数
        ncolumns = 1                                        # 写入列数
        a = 0
        b = 0
        c = 0
        if comb.get() == '高数':
            a = int(entryScore.get())
        elif comb.get == 'C':
            b = int(entryScore.get())
        elif comb.get() == 'Python':
            c = int(entryScore.get())
        values = [int(entryID.get()), '', '', '', a, b, c]  # 需要写入的值
        for value in values:
            table.cell(nrows + 1, ncolumns).value = value   # 每列写入一个值
            ncolumns = ncolumns + 1                         # 列数加1
        data.save('../lib/Students.xlsx')                   # 追加写入保存
        messagebox.showinfo('提示', '提交成功！')

    # 点击"返回"按钮的操作
    def clickReturnR():
        modifyRoot.destroy()

    # 初始化控件
    idRoot = tk.Frame(modifyRoot, bg='lightblue')
    idRoot.pack()
    subjectRoot = tk.Frame(modifyRoot, bg='lightblue')
    subjectRoot.pack()
    scoreRoot = tk.Frame(modifyRoot, bg='lightblue')
    scoreRoot.pack()
    btnRoot = tk.Frame(modifyRoot, bg='lightblue')
    btnRoot.pack()

    # 输入获取
    var_id = tk.StringVar()
    var_subject = tk.StringVar()
    var_score = tk.StringVar()

    # 学号控件
    tk.Label(idRoot, text='学号', bg='lightblue').grid(row=1, column=0, ipady=30, ipadx=10)
    entryID = Entry(idRoot, textvariable=var_id)
    entryID.grid(row=1, column=1, columnspan=3)

    # 科目控件
    tk.Label(subjectRoot, text='科目', bg='lightblue').grid(row=2, column=0, ipady=30, ipadx=10)
    comb = Combobox(subjectRoot, textvariable=var_subject, values=['高数', 'C', 'Python'], width=18)
    comb.grid(row=2, column=1)

    # 成绩控件
    tk.Label(scoreRoot, text='成绩', bg='lightblue').grid(row=3, column=0, ipady=30, ipadx=10)
    entryScore = Entry(scoreRoot, textvariable=var_score)
    entryScore.grid(row=3, column=1, columnspan=3)

    # 按钮显示
    Button(btnRoot, text='提交', width=6, command=clickSubR).grid(row=4, column=0)
    Button(btnRoot, text='返回', width=6, command=clickReturnR).grid(row=4, column=1)

    modifyRoot.mainloop()

if __name__ == '__main__':
    ModifyRoot()