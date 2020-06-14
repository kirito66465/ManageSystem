'''
@coding=UTF-8
@auther:Yan Chen, Fei JianLin, Yang Cheng, Tang YuHao
@time:2020-06
@ui:View Info of Stu
'''

# 导入模块
import tkinter as tk
from tkinter import *
import openpyxl
from src.ui.UserID import *

'''
下面是生成查看我的信息页面的代码
'''
def ViewInfoStu():
    viewInfoStu = tk.Tk()
    viewInfoStu.title('欢迎来到学生管理系统！')
    viewInfoStu.geometry('600x350+300+300')
    viewInfoStu['bg'] = 'lightblue'
    viewInfoStu.attributes('-alpha', 0.9)

    # 读取表格
    def readS():
        ioS = r'../lib/Students.xlsx'                               # 读取Excel表格路径
        data = openpyxl.load_workbook(ioS)                          # 读取Excel表格
        table = data.get_sheet_by_name('Sheet1')                    # 获得指定名称的页
        nrows = table.rows                                          # 获得行数 类型为迭代器
        for row in nrows:
            line = [col.value for col in row]                       # 取值
        n = int(getID())
        showIDS.configure(text=table.cell(n, 1).value)              # 读取第n行即第n-1个学生的学号
        showNameS.configure(text=table.cell(n, 2).value)            # 读取第n行即第n-1个学生的姓名
        showSexS.configure(text=table.cell(n, 3).value)             # 读取第n行即第n-1个学生的性别
        showAgeS.configure(text=table.cell(n, 4).value)             # 读取第n行即第n-1个学生的年龄
        showSubjectS.configure(text='高数、C、Python')

    # 点击"返回"按钮的操作
    def clickReturnS1():
        viewInfoStu.destroy()

    # 初始化控件
    showInfoStu1 = tk.Frame(viewInfoStu)
    showInfoStu1.pack()
    showInfoStu2 = tk.Frame(viewInfoStu)
    showInfoStu2.pack()
    showInfoStu3 = tk.Frame(viewInfoStu)
    showInfoStu3.pack()
    showInfoStu4 = tk.Frame(viewInfoStu)
    showInfoStu4.pack()
    showInfoStu5 = tk.Frame(viewInfoStu)
    showInfoStu5.pack()
    btnStu = tk.Frame(viewInfoStu)
    btnStu.pack()

    # 信息显示
    tk.Label(showInfoStu1, text='学号', bg='lightblue', width=5).grid(row=1, column=0, ipady=7)
    showIDS = tk.Label(showInfoStu1, text='', relief=GROOVE, width=15)
    showIDS.grid(row=1, column=1, ipady=7)

    tk.Label(showInfoStu2, text='姓名', bg='lightblue', width=5).grid(row=2, column=0, ipady=7)
    showNameS = tk.Label(showInfoStu2, text='', relief=GROOVE, width=15)
    showNameS.grid(row=2, column=1, ipady=7)

    tk.Label(showInfoStu3, text='性别', bg='lightblue', width=5).grid(row=3, column=0, ipady=7)
    showSexS = tk.Label(showInfoStu3, text='', relief=GROOVE, width=15)
    showSexS.grid(row=3, column=1, ipady=7)

    tk.Label(showInfoStu4, text='年龄', bg='lightblue', width=5).grid(row=4, column=0, ipady=7)
    showAgeS = tk.Label(showInfoStu4, text='', relief=GROOVE, width=15)
    showAgeS.grid(row=4, column=1, ipady=7)

    tk.Label(showInfoStu5, text='选课', bg='lightblue', width=5).grid(row=5, column=0, ipady=7)
    showSubjectS = tk.Label(showInfoStu5, text='', relief=GROOVE, width=15)
    showSubjectS.grid(row=5, column=1, ipady=7)

    readS()

    # 按钮显示
    tk.Button(btnStu, text='返回', width=10, height=1, relief=GROOVE, command=clickReturnS1).grid(row=6, column=0, ipady=7)

    viewInfoStu.mainloop()

if __name__ == '__main__':
    ViewInfoStu()