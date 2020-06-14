'''
@coding=UTF-8
@auther:Yan Chen, Fei JianLin, Yang Cheng, Tang YuHao
@time:2020-06
@ui:View Score of Stu
'''

# 导入模块
import tkinter as tk
from tkinter import *
import openpyxl
from src.ui.UserID import *

'''
下面是生成查看我的成绩页面的代码
'''
def ViewScoreStu():
    viewScoreStu = tk.Tk()
    viewScoreStu.title('欢迎来到学生管理系统！')
    viewScoreStu.geometry('600x350+300+300')
    viewScoreStu['bg'] = 'lightblue'
    viewScoreStu.attributes('-alpha', 0.9)

    # 读取表格
    def readS():
        ioS = r'../lib/Students.xlsx'                               # 读取Excel表格路径
        data = openpyxl.load_workbook(ioS)                          # 读取Excel表格
        table = data.get_sheet_by_name('Sheet1')                    # 获得指定名称的页
        nrows = table.rows                                          # 获得行数 类型为迭代器
        for row in nrows:
            line = [col.value for col in row]                       # 取值
        n = int(getID())
        showS1.configure(text=table.cell(n, 5).value)               # 读取第n行即第n-1个学生的高数成绩
        showS2.configure(text=table.cell(n, 6).value)               # 读取第n行即第n-1个学生的C成绩
        showS3.configure(text=table.cell(n, 7).value)               # 读取第n行即第n-1个学生的Python成绩

    # 点击"返回"按钮的操作
    def clickReturnS2():
        viewScoreStu.destroy()

    # 初始化控件
    showScoreStu1 = tk.Frame(viewScoreStu)
    showScoreStu1.pack()
    showScoreStu2 = tk.Frame(viewScoreStu)
    showScoreStu2.pack()
    showScoreStu3 = tk.Frame(viewScoreStu)
    showScoreStu3.pack()
    btnStu = tk.Frame(viewScoreStu)
    btnStu.pack()

    # 信息显示
    tk.Label(showScoreStu1, text='高数', bg='lightblue', width=5).grid(row=1, column=0, ipady=7)
    showS1 = tk.Label(showScoreStu1, text='', relief=GROOVE, width=15)
    showS1.grid(row=1, column=1, ipady=7)

    tk.Label(showScoreStu2, text='C', bg='lightblue', width=5).grid(row=2, column=0, ipady=7)
    showS2 = tk.Label(showScoreStu2, text='', relief=GROOVE, width=15)
    showS2.grid(row=2, column=1, ipady=7)

    tk.Label(showScoreStu3, text='Python', bg='lightblue', width=5).grid(row=3, column=0, ipady=7)
    showS3 = tk.Label(showScoreStu3, text='', relief=GROOVE, width=15)
    showS3.grid(row=3, column=1, ipady=7)

    readS()

    # 按钮显示
    tk.Button(btnStu, text='返回', width=10, height=1, relief=GROOVE, command=clickReturnS2).grid(row=4, column=0, ipady=7)

    viewScoreStu.mainloop()

if __name__ == '__main__':
    ViewScoreStu()